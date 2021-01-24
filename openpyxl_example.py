# 引入基础的工作表
from openpyxl import Workbook 
# 引入增强的修改功能
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill,colors
# import openpyxl
def make_example():
    # 设定文件目录
    addr = './example.xlsx'
    # 初始化文件，切换到活动的工作表
    work_book = Workbook()
    # 读取文件采用
    # work_book = openpyxl.load_workbook(addr)
    work_sheet = work_book.active
    # 直接对表格对象赋值
    work_sheet['A1'] = 'Hello World!'
    # 采用指定行列的方法赋值(第2行，第二列）
    select_cell = work_sheet.cell(row=2,column=2,value='I select this cell')
    # 添加两行数据到表格
    work_sheet.append(['The quick brown fox',' jumps over ','a lazy dog.'])
    work_sheet.append(['The quick brown fox',' ',' jumps over ','a lazy dog.'])
    # 合并两个单元格作为示范
    work_sheet.merge_cells('A3:B3')
    work_sheet.merge_cells('A4:B4')
    # 遍历表格，读取表格中的数据
    # 初始化字体
    SIMSUN_20_BOLD = Font(name='宋体',size=12,bold=True)
    # 初始化表格对齐模板
    CENTER_ALIGN = Alignment(horizontal='center',vertical='center')
    # 初始化表格边框样式
    LE,RI,TO,BO = [Side(style='thin',color='000000')]*4
    THIN_BORDER = Border(left=LE,right=RI,top=TO,bottom=BO)
    # 遍历表格，读取表格中的数据
    for row in work_sheet['A1:D4']:
        for cell in row:
            # 把样式赋值给表格
            cell.font = SIMSUN_20_BOLD
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            # print(cell.value)
    # 设置行高
    work_sheet.row_dimensions[1].height=15
    work_sheet.row_dimensions[2].height=20
    for row_letter in range(3,5,1):
        work_sheet.row_dimensions[row_letter].height=17
    # 设置列宽
    for col_letter in ['A','B']:
        work_sheet.column_dimensions[col_letter].width=20
    work_sheet.column_dimensions['C'].width=17
    work_sheet.column_dimensions['D'].width=25
    # 设置颜色
    COLOR_MAP = ['ff9900','000000']
    COLOR_SIMSUN_20_BOLD = Font(name='宋体',size=12,bold=True,color=COLOR_MAP[0])
    BG_FILL = PatternFill('solid', fgColor=COLOR_MAP[1]) 
    work_sheet['A1'].font = COLOR_SIMSUN_20_BOLD
    work_sheet['A1'].fill = BG_FILL
    # 保存到设定的addr
    work_book.save(addr)

if __name__ == "__main__":
    make_example()